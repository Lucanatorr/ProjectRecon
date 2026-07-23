"""Excel/PDF report builders.

The Excel workbook is the reviewable, sign-off-able artifact and audit trail:
tabs for Summary, Flagged, Full detail, Unmatched, with conditional formatting on
variance. See spec §9 / SDD §5.8.
"""
from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors

from recon.models import ReconRow, Severity
from recon.reconcile import CycleTotals

# Severity → fill color (matches the mockup palette).
_FILL = {
    Severity.CRITICAL: PatternFill("solid", fgColor="FBE9E7"),
    Severity.WARNING: PatternFill("solid", fgColor="FBF0DF"),
    Severity.INFO: PatternFill("solid", fgColor="EEF1F6"),
    Severity.OK: PatternFill("solid", fgColor="E6F2EC"),
}
_HEADER_FILL = PatternFill("solid", fgColor="18223A")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
_MONEY = "$#,##0.00"
_NUM = "#,##0.###"

_DETAIL_COLS = [
    ("Code", "code"), ("Description", "description"), ("UoM", "uom"),
    ("Built qty", "built_qty"), ("Billed qty", "billed_qty"),
    ("Contract $", "contract_price"), ("Billed $", "billed_price"),
    ("Est qty", "est_qty"), ("Billed amount", "billed_amount"),
    ("Expected amount", "expected_amount"), ("Variance", "amount_variance"),
    ("Severity", "severity"), ("Flags", "flags"),
]


def _row_values(r: ReconRow) -> dict:
    return {
        "code": r.code or "—",
        "description": r.description,
        "uom": r.uom.value,
        "built_qty": r.built_qty,
        "billed_qty": r.billed_qty,
        "contract_price": r.contract_price,
        "billed_price": r.billed_price,
        "est_qty": r.est_qty,
        "billed_amount": r.billed_amount,
        "expected_amount": r.expected_amount,
        "amount_variance": r.amount_variance,
        "severity": r.severity.value,
        "flags": " | ".join(f.message for f in r.flags),
    }


def _write_sheet(ws, rows: list[ReconRow]) -> None:
    # header
    for c, (label, _) in enumerate(_DETAIL_COLS, start=1):
        cell = ws.cell(row=1, column=c, value=label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left")
    # body
    for ri, r in enumerate(rows, start=2):
        vals = _row_values(r)
        fill = _FILL.get(r.severity)
        for ci, (_, key) in enumerate(_DETAIL_COLS, start=1):
            cell = ws.cell(row=ri, column=ci, value=vals[key])
            if key in ("contract_price", "billed_price", "billed_amount",
                       "expected_amount", "amount_variance"):
                cell.number_format = _MONEY
            elif key in ("built_qty", "billed_qty", "est_qty"):
                cell.number_format = _NUM
            if fill and key in ("severity", "amount_variance"):
                cell.fill = fill
    # widths
    for c, (label, _) in enumerate(_DETAIL_COLS, start=1):
        ws.column_dimensions[get_column_letter(c)].width = max(12, len(label) + 4)
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["M"].width = 50
    ws.freeze_panes = "A2"


def _write_summary(ws, totals: CycleTotals, cycle_label: str,
                  override: dict | None = None) -> None:
    ws["A1"] = "Splice — Reconciliation Summary"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = cycle_label
    ws["A2"].font = Font(color="5C6B80")

    metrics = [
        ("Billed this cycle (gross)", totals.total_billed, _MONEY),
        ("Expected (built × contract)", totals.total_expected, _MONEY),
        ("Flagged over-billing", totals.flagged_over_billing, _MONEY),
        ("Retainage held", totals.retainage_held, _MONEY),
        ("Net recommended", totals.net_recommended, _MONEY),
        ("Critical flags", totals.n_critical, "0"),
        ("Warnings", totals.n_warning, "0"),
        ("Reconciled / info", totals.n_ok, "0"),
    ]
    for i, (label, val, fmt) in enumerate(metrics, start=4):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        c = ws.cell(row=i, column=2, value=val)
        c.number_format = fmt
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18

    if override:
        row = 4 + len(metrics) + 1
        who = override.get("by") or "reviewer"
        cell = ws.cell(row=row, column=1,
                      value="EXPORTED WITH OVERRIDE — pre-export checks did not pass")
        cell.font = Font(bold=True, color="C6362F")
        ws.cell(row=row + 1, column=1,
                value=f"Reason: {override.get('reason', '—')} ({who} "
                      f"{override.get('at', '')})").font = Font(color="5C6B80")


def build_workbook(rows: list[ReconRow], totals: CycleTotals,
                  cycle_label: str = "", *, override: dict | None = None) -> bytes:
    """Build the multi-tab Excel workbook and return it as bytes."""
    wb = Workbook()

    ws_sum = wb.active
    ws_sum.title = "Summary"
    _write_summary(ws_sum, totals, cycle_label, override)

    flagged = [r for r in rows if r.severity in (Severity.CRITICAL, Severity.WARNING)]
    _write_sheet(wb.create_sheet("Flagged"), flagged)
    _write_sheet(wb.create_sheet("Full detail"), rows)

    unmatched = [r for r in rows if r.code is None or r.contract_price is None]
    _write_sheet(wb.create_sheet("Unmatched"), unmatched)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------- #
#  PDF approval summary
# --------------------------------------------------------------------------- #
# Palette mirrored from the UI so the packet looks like the app.
_INK = colors.HexColor("#101826")
_BLUE = colors.HexColor("#1c5ac4")
_CRITICAL = colors.HexColor("#c6362f")
_WARN = colors.HexColor("#c9781a")
_MUTED = colors.HexColor("#5c6b80")
_LINE = colors.HexColor("#d7dee8")
_CANVAS = colors.HexColor("#eef1f6")

_SEV_COLOR = {Severity.CRITICAL: _CRITICAL, Severity.WARNING: _WARN}


def _money(v: float) -> str:
    sign = "-" if v < 0 else ""
    return f"{sign}${abs(v):,.2f}"


def build_pdf_summary(rows: list[ReconRow], totals: CycleTotals,
                     cycle_label: str = "", *, max_flagged: int = 18,
                     resolutions: dict | None = None,
                     override: dict | None = None) -> bytes:
    """One-page PDF summary for the payment-approval packet.

    Headline metrics, the payment recommendation, the flagged items holding money
    back, and a sign-off block. Returns PDF bytes.
    """
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        KeepTogether, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    base = getSampleStyleSheet()
    st_title = ParagraphStyle("t", parent=base["Title"], fontSize=16, leading=19,
                              alignment=0, textColor=_INK, spaceAfter=2)
    st_sub = ParagraphStyle("s", parent=base["Normal"], fontSize=9.5,
                            textColor=_MUTED, spaceAfter=2)
    st_h = ParagraphStyle("h", parent=base["Normal"], fontSize=10.5, leading=13,
                          textColor=_INK, spaceBefore=12, spaceAfter=5,
                          fontName="Helvetica-Bold")
    st_cell = ParagraphStyle("c", parent=base["Normal"], fontSize=7.6, leading=9.2)
    st_note = ParagraphStyle("n", parent=base["Normal"], fontSize=9, leading=12,
                             textColor=_INK)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=letter, title="Reconciliation Summary",
        leftMargin=0.6 * inch, rightMargin=0.6 * inch,
        topMargin=0.55 * inch, bottomMargin=0.55 * inch)
    flow = []

    # --- header ---
    flow.append(Paragraph("Splice — Reconciliation Summary", st_title))
    if cycle_label:
        flow.append(Paragraph(cycle_label, st_sub))
    flow.append(Paragraph(
        f"Generated {datetime.now():%Y-%m-%d %H:%M} · contractor invoice reconciled "
        "against documented as-built quantities", st_sub))

    # An overridden export must say so on the artifact itself.
    if override:
        st_ovr = ParagraphStyle("o", parent=base["Normal"], fontSize=9, leading=12,
                                textColor=colors.white, backColor=_CRITICAL,
                                borderPadding=5, spaceBefore=8)
        who = override.get("by") or "reviewer"
        when = override.get("at") or ""
        flow.append(Paragraph(
            f"<b>EXPORTED WITH OVERRIDE</b> — pre-export checks did not pass. "
            f"Reason: {override.get('reason', '—')} ({who} {when})", st_ovr))
    flow.append(Spacer(1, 10))

    # --- headline metrics ---
    metrics = [
        ("Billed this cycle (gross)", _money(totals.total_billed), _INK),
        ("Expected (built × contract)", _money(totals.total_expected), _INK),
        ("Flagged over-billing", _money(totals.flagged_over_billing), _CRITICAL),
        ("Retainage held", _money(totals.retainage_held), _INK),
        ("Net recommended", _money(totals.net_recommended), _INK),
    ]
    mt = Table([[k, v] for k, v, _ in metrics], colWidths=[3.4 * inch, 3.9 * inch])
    style = [
        ("FONTSIZE", (0, 0), (-1, -1), 9.5),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("FONTNAME", (1, 0), (1, -1), "Helvetica-Bold"),
        ("LINEBELOW", (0, 0), (-1, -2), 0.4, _LINE),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("BACKGROUND", (0, len(metrics) - 1), (-1, len(metrics) - 1), _CANVAS),
    ]
    for i, (_, _, col) in enumerate(metrics):
        style.append(("TEXTCOLOR", (1, i), (1, i), col))
    mt.setStyle(TableStyle(style))
    flow.append(mt)
    flow.append(Spacer(1, 8))

    flow.append(Paragraph(
        f"<b>Recommended payment this cycle: {_money(totals.net_recommended)}</b> — "
        f"holds {_money(totals.flagged_over_billing)} in flagged items and withholds "
        f"{_money(totals.retainage_held)} retainage. "
        f"{totals.n_critical} critical · {totals.n_warning} warning · "
        f"{totals.n_ok} reconciled.", st_note))

    # --- flagged items ---
    flagged = [r for r in rows
               if r.severity in (Severity.CRITICAL, Severity.WARNING)]
    flow.append(Paragraph(
        f"Flagged items ({len(flagged)}) — reviewed before payment", st_h))

    if not flagged:
        flow.append(Paragraph("No flagged items. Everything reconciled cleanly.",
                              st_note))
    else:
        head = ["Code", "Unit", "Built", "Billed", "Variance", "Severity", "Finding"]
        if resolutions:
            head.append("Resolution")
        data = [head]
        shown = flagged[:max_flagged]
        for r in shown:
            finding = "; ".join(f.message for f in r.flags) or "—"
            row_cells = [
                Paragraph(r.code or "—", st_cell),
                Paragraph(r.description, st_cell),
                Paragraph(f"{r.built_qty:,.0f}", st_cell),
                Paragraph(f"{r.billed_qty:,.0f}", st_cell),
                Paragraph(_money(r.amount_variance), st_cell),
                Paragraph(r.severity.value.upper(), st_cell),
                Paragraph(finding, st_cell),
            ]
            if resolutions:
                res = resolutions.get(r.code or r.description) or {}
                label = (res.get("status") or "open").upper()
                if res.get("note"):
                    label += f" — {res['note']}"
                row_cells.append(Paragraph(label, st_cell))
            data.append(row_cells)

        # widths sum to the 7.3in printable width (letter minus 0.6in margins)
        widths = ([0.45, 1.45, 0.55, 0.55, 0.75, 0.80, 2.75] if not resolutions
                  else [0.42, 1.15, 0.48, 0.48, 0.68, 0.80, 1.85, 1.44])
        ft = Table(data, colWidths=[w * inch for w in widths], repeatRows=1)
        ts = [
            ("BACKGROUND", (0, 0), (-1, 0), _INK),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 7.6),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.3, _LINE),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]
        for i, r in enumerate(shown, start=1):
            col = _SEV_COLOR.get(r.severity, _MUTED)
            ts.append(("TEXTCOLOR", (5, i), (5, i), col))
            ts.append(("TEXTCOLOR", (4, i), (4, i),
                       _CRITICAL if r.amount_variance > 0.5 else _MUTED))
        ft.setStyle(TableStyle(ts))
        flow.append(ft)
        if len(flagged) > max_flagged:
            flow.append(Spacer(1, 4))
            flow.append(Paragraph(
                f"+ {len(flagged) - max_flagged} more flagged item(s) — see the Excel "
                "workbook for the full detail.", st_sub))

    # --- sign-off ---
    flow.append(Paragraph("Sign-off", st_h))
    sign = Table(
        [["Reviewed by", "", "Date", ""], ["Approved by", "", "Date", ""]],
        colWidths=[0.9 * inch, 3.3 * inch, 0.5 * inch, 1.6 * inch])
    sign.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TEXTCOLOR", (0, 0), (0, -1), _MUTED),
        ("TEXTCOLOR", (2, 0), (2, -1), _MUTED),
        ("LINEBELOW", (1, 0), (1, -1), 0.6, _INK),
        ("LINEBELOW", (3, 0), (3, -1), 0.6, _INK),
        ("TOPPADDING", (0, 0), (-1, -1), 12),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    flow.append(KeepTogether(sign))
    flow.append(Spacer(1, 6))
    flow.append(Paragraph(
        "This tool produces a payment recommendation from documented as-built "
        "quantities and the contract bid schedule; it does not disburse funds.",
        st_sub))

    doc.build(flow)
    return buf.getvalue()
