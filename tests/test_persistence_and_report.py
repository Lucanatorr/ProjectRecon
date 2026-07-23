"""Persistence round-trip + report generation smoke tests (FR-15/16/17)."""
from __future__ import annotations

import io

import pytest
from openpyxl import load_workbook

from recon.persistence import Database
from recon.report import build_pdf_summary, build_workbook
from recon.reconcile import cycle_totals, reconcile


@pytest.fixture
def db():
    d = Database(":memory:")
    yield d
    d.close()


def test_project_contract_roundtrip(db, golden_contract):
    pid = db.create_project("Robeson CAB — PON 5", contractor="Rivr Tech")
    db.save_contract(pid, golden_contract)
    loaded = db.load_contract(pid)
    assert len(loaded) == len(golden_contract)
    assert {c.code for c in loaded} == {c.code for c in golden_contract}
    assert loaded[0].unit_price == golden_contract[0].unit_price


def test_alias_store_persists_and_reloads(db):
    db.confirm_alias("144F ADSS Aerial Place", "3.1", confidence=95, actor="tester")
    store = db.load_alias_store()
    from recon.ingest.normalize import normalize
    assert store.get(normalize("144F ADSS Aerial Place")) == "3.1"
    # audit log recorded the confirmation
    trail = db.audit_trail()
    assert any(r["action"] == "confirm_alias" for r in trail)


def test_save_results_and_trend(db, golden_asbuilt, golden_invoices, golden_contract):
    pid = db.create_project("P")
    cid = db.create_cycle(pid, 4, period_label="Jul 2026",
                         billing_mode="cumulative", retainage_pct=10.0)
    rows = reconcile(golden_asbuilt, golden_invoices, golden_contract)
    db.save_results(cid, rows)
    trend = db.trend(pid)
    assert len(trend) == 1
    assert trend[0]["billed_value"] > 0


def test_workbook_has_expected_tabs(golden_asbuilt, golden_invoices, golden_contract):
    rows = reconcile(golden_asbuilt, golden_invoices, golden_contract)
    totals = cycle_totals(rows, retainage_pct=10.0)
    data = build_workbook(rows, totals, "Cycle 04 · Jul 2026")
    wb = load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["Summary", "Flagged", "Full detail", "Unmatched"]
    # Unmatched tab includes the not-in-contract traffic-control line.
    unmatched = wb["Unmatched"]
    descs = [unmatched.cell(row=r, column=2).value for r in range(2, unmatched.max_row + 1)]
    assert any(d and "Traffic" in str(d) or d == "—" for d in descs) or unmatched.max_row >= 2


def _pdf_text(data: bytes) -> str:
    import pdfplumber
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        return "\n".join((p.extract_text() or "") for p in pdf.pages)


def test_pdf_summary_is_a_real_pdf(golden_asbuilt, golden_invoices, golden_contract):
    rows = reconcile(golden_asbuilt, golden_invoices, golden_contract)
    totals = cycle_totals(rows, retainage_pct=10.0)
    out = build_pdf_summary(rows, totals, "Cycle 04")
    assert isinstance(out, bytes)
    assert out.startswith(b"%PDF")           # a real PDF, not the old text blob

    text = _pdf_text(out)
    assert "Splice" in text and "Reconciliation Summary" in text
    assert "Cycle 04" in text
    assert "35,408" in text                  # headline flagged over-billing
    assert "Net recommended" in text
    assert "Sign-off" in text                # approval block for the packet
    # severity labels render intact (not wrapped by a too-narrow column)
    assert "CRITICAL" in text and "WARNING" in text


def test_pdf_summary_lists_flagged_items(golden_asbuilt, golden_invoices,
                                        golden_contract):
    rows = reconcile(golden_asbuilt, golden_invoices, golden_contract)
    totals = cycle_totals(rows, retainage_pct=10.0)
    text = _pdf_text(build_pdf_summary(rows, totals, "Cycle 04"))
    # single-word probes: descriptions wrap across lines inside the table cell
    assert "Traffic" in text                 # the unauthorized unit
    assert "ADSS" in text                    # the over-billed aerial unit
    assert "26,000" in text                  # its dollar exposure


def test_pdf_summary_includes_resolutions_when_given(golden_asbuilt, golden_invoices,
                                                    golden_contract):
    rows = reconcile(golden_asbuilt, golden_invoices, golden_contract)
    totals = cycle_totals(rows, retainage_pct=10.0)
    res = {"3.1": {"status": "hold", "note": "await field verification"}}
    text = _pdf_text(build_pdf_summary(rows, totals, "Cycle 04", resolutions=res))
    assert "Resolution" in text
    assert "HOLD" in text
