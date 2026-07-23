"""Sprint 4.3 — pre-export validation gates + logged override (SDD §9)."""
from __future__ import annotations

import io

import pytest

from recon.models import AsBuiltLine, ContractItem, InvoiceLine, UoM
from recon.reconcile import cycle_totals, reconcile
from recon.report import build_pdf_summary, build_workbook
from ui.gates import blocking, evaluate_gates
from ui.state import WizardState, row_key, set_resolution


def _clean_state() -> WizardState:
    """A cycle where every gate passes."""
    contract = [ContractItem("A", "Unit A", UoM.EA, 10.0, 100)]
    ab = [AsBuiltLine("A", 10, UoM.EA, code="A", confidence="sum")]
    iv = [InvoiceLine("INV", "A", 10, 10.0, 100.0, code="A")]
    s = WizardState()
    s.contract = contract
    s.asbuilt = ab
    s.invoices = iv
    s.resolved = {"A": "A"}                     # crosswalk settled
    s.results = reconcile(ab, iv, contract)
    return s


def _gate(state, key):
    return next(g for g in evaluate_gates(state) if g.key == key)


def test_clean_cycle_passes_every_gate():
    s = _clean_state()
    gates = evaluate_gates(s)
    assert all(g.ok for g in gates), [g.key for g in gates if not g.ok]
    assert blocking(gates) == []


def test_missing_bid_schedule_blocks():
    s = _clean_state()
    s.contract = []
    assert _gate(s, "contract").ok is False
    assert "contract" in {g.key for g in blocking(evaluate_gates(s))}


def test_unconfirmed_extracted_rows_block():
    s = _clean_state()
    s.asbuilt = [AsBuiltLine("A", 10, UoM.EA, code="A", confidence="pdf")]
    g = _gate(s, "confidence")
    assert g.ok is False
    assert "not yet confirmed" in g.detail
    # confirming them clears the gate
    s.asbuilt[0].confidence = "confirmed"
    assert _gate(s, "confidence").ok is True


def test_unresolved_crosswalk_blocks():
    s = _clean_state()
    # a description that can't auto-map and hasn't been confirmed
    s.invoices = s.invoices + [InvoiceLine("INV", "Mystery widget", 1, 5.0, 5.0)]
    assert _gate(s, "crosswalk").ok is False


def test_unresolved_criticals_block_and_clear_on_decision():
    contract = [ContractItem("A", "Unit A", UoM.EA, 10.0, 100)]
    ab = [AsBuiltLine("A", 10, UoM.EA, code="A")]
    iv = [InvoiceLine("INV", "A", 50, 10.0, 500.0, code="A")]     # qty over -> critical
    s = WizardState()
    s.contract, s.asbuilt, s.invoices = contract, ab, iv
    s.resolved = {"A": "A"}
    s.results = reconcile(ab, iv, contract)

    assert _gate(s, "criticals").ok is False
    for r in s.results:
        if r.severity.value == "critical":
            set_resolution(s, row_key(r), "hold", by="LC")
    assert _gate(s, "criticals").ok is True


def test_gates_report_the_step_that_fixes_them():
    s = _clean_state()
    steps = {g.key: g.fix_step for g in evaluate_gates(s)}
    assert steps == {"contract": "contract", "confidence": "asbuilt",
                     "crosswalk": "crosswalk", "criticals": "reconcile"}


# --- override is stamped onto the artifacts ---
def _totals_and_rows():
    contract = [ContractItem("A", "Unit A", UoM.EA, 10.0, 100)]
    rows = reconcile([AsBuiltLine("A", 10, UoM.EA, code="A")],
                     [InvoiceLine("INV", "A", 50, 10.0, 500.0, code="A")], contract)
    return rows, cycle_totals(rows, retainage_pct=10.0)


def test_pdf_stamps_the_override():
    rows, totals = _totals_and_rows()
    ovr = {"reason": "paying uncontested items now", "by": "LC", "at": "2026-07-23 18:00"}
    pdf = build_pdf_summary(rows, totals, "Cycle 04", override=ovr)

    import pdfplumber
    with pdfplumber.open(io.BytesIO(pdf)) as d:
        text = "\n".join((p.extract_text() or "") for p in d.pages)
    assert "EXPORTED WITH OVERRIDE" in text
    assert "paying uncontested items now" in text


def test_pdf_has_no_override_banner_when_clean():
    rows, totals = _totals_and_rows()
    import pdfplumber
    with pdfplumber.open(io.BytesIO(build_pdf_summary(rows, totals, "Cycle 04"))) as d:
        text = "\n".join((p.extract_text() or "") for p in d.pages)
    assert "OVERRIDE" not in text


def test_workbook_stamps_the_override():
    from openpyxl import load_workbook
    rows, totals = _totals_and_rows()
    ovr = {"reason": "closeout deadline", "by": "LC", "at": "2026-07-23 18:00"}
    wb = load_workbook(io.BytesIO(build_workbook(rows, totals, "Cycle 04", override=ovr)))
    text = "\n".join(str(c.value) for row in wb["Summary"].iter_rows()
                     for c in row if c.value)
    assert "EXPORTED WITH OVERRIDE" in text
    assert "closeout deadline" in text
